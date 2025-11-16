from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QFileDialog
from openpyxl import Workbook


def exportToExcel(model):
    if model is None:
        return

    filePath, _ = QFileDialog.getSaveFileName(
        None,
        "Сохранить как",
        "",
        "Excel (*.xlsx)"
    )

    if not filePath:
        return

    wb = Workbook()
    ws = wb.active

    # Заголовоки
    for col in range(model.columnCount()):
        header = model.headerData(col, Qt.Orientation.Horizontal)
        ws.cell(row=1, column=col+1, value=header)

    # Строки
    for row in range(model.rowCount()):
        for col in range(model.columnCount()):
            index = model.index(row, col)
            ws.cell(row=row + 2, column=col + 1, value=str(model.data(index)))

    wb.save(filePath)